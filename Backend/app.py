import os
import io
import pandas as pd
from flask import session
from flask_cors import CORS
from flask_mail import Mail, Message
from openpyxl.styles import Border, Side
from Backend.config import app, db, db_path
from openpyxl.utils import get_column_letter
from Backend.models import Course, User, UpdateRequest, DatabaseHandler
from flask import send_file, request, render_template, redirect, url_for, jsonify


CORS(app)  # Enable CORS for all routes

# Configure Flask-Mail
app.config['MAIL_SERVER'] = os.getenv('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.getenv('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME', 'patelshyamal16@gmail.com')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD', 'eplt yzca osub mady')
mail = Mail(app)

@app.route("/", methods=["GET", "POST"])
def home():
    if request.method == "POST":
        password = request.form.get("password")
        # You can change this password or get it from environment variable
        correct_password = os.environ.get("DASHBOARD_PASSWORD", "admin123")
        if password == correct_password:
            session['logged_in'] = True
            return redirect(url_for('home'))
        else:
            error = "Invalid password. Please try again."
            return render_template("login.html", error=error, login_type='dashboard')
    else:
        if session.get('logged_in'):
            users = User.query.all()
            return render_template("dashboard.html", members=users)
        else:
            return render_template("login.html", login_type='dashboard')

@app.route("/add_member")
def add_member():
    return render_template("user.html")

@app.route("/submit_user", methods=["POST"])
def submit_user():
    first_name = request.form.get("first_name")
    last_name = request.form.get("last_name")
    email = request.form.get("email")
    clinical_appe = int(request.form.get("clinical_appe", 0))
    academic_appe = int(request.form.get("academic_appe", 0))
    research_electives = int(request.form.get("research_electives", 0))
    proctoring = float(request.form.get("proctoring", 0))
    grading =  int(request.form.get("grading", 0))

    if clinical_appe > 0 and academic_appe > 0:
        total_appe = (2.75 * clinical_appe) + (1.4 * academic_appe)
    elif clinical_appe > 0 and academic_appe == 0:
        total_appe = (2.75 * clinical_appe)
    else:
        total_appe = (1.4 * academic_appe)

    user = User.query.filter_by(email=email).first()
    if not user:
        user = User(
            first_name=first_name,
            last_name=last_name,
            email=email,
            clinical_appe=clinical_appe,
            academic_appe=academic_appe,
            research_electives=research_electives,
            proctoring=proctoring,
            grading=grading,
            total_appe=total_appe
        )
        db.session.add(user)
    else:
        user.first_name = first_name
        user.last_name = last_name
        user.clinical_appe = clinical_appe
        user.academic_appe = academic_appe
        user.research_electives = research_electives
        user.proctoring = proctoring
        user.grading = grading
        user.total_appe = total_appe
    db.session.commit()

    return "", 200

@app.route("/submit_course", methods=["POST"])
def submit_course():
    user_id = request.form.get("user_id")
    if not user_id:
        return "User ID not provided", 400
    user = User.query.filter_by(user_id=int(user_id)).first()
    if not user:
        return "User not found", 400

    course_name = request.form.get("course_name")
    enroll = request.form.get("enroll", 'N')  # 'Y' or 'N'
    didactic_credit = float(request.form.get("didactic_credit", 0))
    lab_credit = float(request.form.get("lab_credit", 0))
    coordinator = request.form.get("coordinator", 'N')  # 'Y' or 'N'
    clinical_lead = request.form.get("clinical_lead", 'N')  # 'Y' or 'N' or 'NA'
    lecture_total = int(request.form.get("lecture_total") or 0)
    lab_total = float(request.form.get("lab_total") or 0)
    lecture_faculty = int(request.form.get("lecture_faculty") or 0)
    lab_design = float(request.form.get("lab_design") or 0)
    lab_proctor = float(request.form.get("lab_proctor") or 0)

    if enroll and enroll.upper() == "Y":
        part1 = didactic_credit if coordinator and coordinator.upper() == "Y" else 0
    else:
        part1 = 0

    if coordinator and coordinator.upper() == "Y":
        part2 = lab_credit if clinical_lead and clinical_lead.upper() == "NA" else 0
    else:
        part2 = 0

    part3 = lab_credit if clinical_lead and clinical_lead.upper() == "Y" else 0
    part4 = (lecture_faculty / lecture_total) * didactic_credit if lecture_total > 0 else 0
    part5 = (((lab_design + lab_proctor) / lab_total) * lab_credit) if lab_total > 0 else 0

    total = part1 + part2 + part3 + part4 + part5
    total = round(total, 5)

    course = Course(
        course_name =course_name.capitalize(),
        enroll=enroll.upper() if enroll else None,
        didactic_credit=didactic_credit,
        lab_credit=lab_credit,
        coordinator=coordinator.upper() if coordinator else None,
        clinical_lead=clinical_lead.upper() if clinical_lead else None,
        lecture_total=lecture_total,
        lab_total=lab_total,
        lecture_faculty=lecture_faculty,
        lab_design=lab_design,
        lab_proctor=lab_proctor,
        total=total,
        user_id=user.user_id
    )
    db.session.add(course)
    db.session.commit()

    return redirect(url_for('user_info', user_id=user.user_id))

@app.route("/user_info")
def user_info():
    user_id = request.args.get("user_id")
    if not user_id:
        return "User ID not provided", 400

    user = User.query.filter_by(user_id=user_id).first()
    if not user:
        return "User not found", 404

    user_courses = Course.query.filter_by(user_id=user.user_id).all()

    sum_of_courses = sum(course.total for course in user_courses)

    Total = sum_of_courses + (user.total_appe or 0) + (user.research_electives or 0) + (user.proctoring or 0) + (user.grading or 0)

    Total = round(Total, 2)

    Percentage = (Total* (0.7/21)) * 100

    Percentage = int(round(Percentage,0))

    return render_template("user_info.html", user=user, user_courses=user_courses, sum_of_courses=sum_of_courses, Total=Total, Percent = Percentage)

@app.route("/update_course/<int:course_id>", methods=["POST"])
def update_course(course_id):
    course = Course.query.filter_by(course_id=course_id).first()
    if not course:
        return {"error": "Course not found"}, 404

    data = request.json
    # Update course fields from data
    course.course_name = data.get("course_name", course.course_name)
    course.enroll = data.get("enroll", course.enroll)
    course.didactic_credit = float(data.get("didactic_credit", course.didactic_credit))
    course.lab_credit = float(data.get("lab_credit", course.lab_credit))
    course.coordinator = data.get("coordinator", course.coordinator)
    course.clinical_lead = data.get("clinical_lead", course.clinical_lead)
    course.lecture_total = int(data.get("lecture_total", course.lecture_total))
    course.lab_total = float(data.get("lab_total", course.lab_total))
    course.lecture_faculty = int(data.get("lecture_faculty", course.lecture_faculty))
    course.lab_design = float(data.get("lab_design", course.lab_design))
    course.lab_proctor = float(data.get("lab_proctor", course.lab_proctor))

    # Recalculate total based on the same logic as in submit_course
    enroll = course.enroll
    didactic_credit = course.didactic_credit
    lab_credit = course.lab_credit
    coordinator = course.coordinator
    clinical_lead = course.clinical_lead
    lecture_total = course.lecture_total
    lab_total = course.lab_total
    lecture_faculty = course.lecture_faculty
    lab_design = course.lab_design
    lab_proctor = course.lab_proctor

    if enroll and enroll.upper() == "Y":
        part1 = didactic_credit if coordinator and coordinator.upper() == "Y" else 0
    else:
        part1 = 0

    if coordinator and coordinator.upper() == "Y":
        part2 = lab_credit if clinical_lead and clinical_lead.upper() == "NA" else 0
    else:
        part2 = 0

    part3 = lab_credit if clinical_lead and clinical_lead.upper() == "Y" else 0
    part4 = (lecture_faculty / lecture_total) * didactic_credit if lecture_total > 0 else 0
    part5 = (((lab_design + lab_proctor) / lab_total) * lab_credit) if lab_total > 0 else 0

    total = part1 + part2 + part3 + part4 + part5
    total = round(total, 5)
    course.total = total

    db.session.commit()

    # After updating course, recalculate sum_of_courses and Total for the user
    user = User.query.filter_by(user_id=course.user_id).first()
    user_courses = Course.query.filter_by(user_id=user.user_id).all()
    sum_of_courses = sum(c.total for c in user_courses)
    Total = sum_of_courses + (user.total_appe or 0) + (user.research_electives or 0) + (user.proctoring or 0) + (user.grading or 0)
    Total = round(Total, 2)
    Percentage = (Total* (0.7/21)) * 100
    Percentage = int(round(Percentage,0))

    return {
        "message": "Course updated successfully",
        "sum_of_courses": sum_of_courses,
        "Total": Total,
        "Percent": Percentage,
        "course_total": total
    }

@app.route("/delete_course/<int:course_id>", methods=["DELETE"])
def delete_course(course_id):
    course = Course.query.filter_by(course_id=course_id).first()
    if not course:
        return {"error": "Course not found"}, 404

    user_id = course.user_id
    db.session.delete(course)
    db.session.commit()

    # After deleting course, recalculate sum_of_courses and Total for the user
    user = User.query.filter_by(user_id=user_id).first()
    user_courses = Course.query.filter_by(user_id=user.user_id).all()
    sum_of_courses = sum(c.total for c in user_courses)
    Total = sum_of_courses + (user.total_appe or 0) + (user.research_electives or 0) + (user.proctoring or 0) + (user.grading or 0)
    Total = round(Total, 2)
    Percentage = (Total* (0.7/21)) * 100
    Percentage = int(round(Percentage,0))

    return {
        "message": "Course deleted successfully",
        "sum_of_courses": sum_of_courses,
        "Total": Total,
        "Percent": Percentage
    }

@app.route("/add_course")
def add_course():
    user_id = request.args.get("user_id")
    if not user_id:
        return "User ID not provided", 400
    return render_template("course.html", user_id=user_id)

@app.route("/edit_user")
def edit_user():
    user_id = request.args.get("user_id")
    user = User.query.get(user_id)
    if not user:
        return "User not found", 404
    return render_template('user_edit.html', user=user)

@app.route('/update_user/<int:user_id>', methods=['POST'])
def update_user(user_id):
    user = User.query.get(user_id)
    if not user:
        return "User not found", 404
    user.first_name = request.form.get('first_name')
    user.last_name = request.form.get('last_name')
    user.email = request.form.get('email')
    clinical_appe = int(request.form.get('clinical_appe', 0))
    user.clinical_appe = clinical_appe
    academic_appe = int(request.form.get('academic_appe', 0))
    user.academic_appe = academic_appe
    user.research_electives = int(request.form.get('research_electives', 0))
    user.proctoring = float(request.form.get('proctoring', 0))
    user.grading = int(request.form.get('grading', 0))


    if clinical_appe > 0 and academic_appe > 0:
        total_appe = (2.75 * clinical_appe) + (1.4 * academic_appe)
    elif clinical_appe > 0 and academic_appe == 0:
        total_appe = (2.75 * clinical_appe)
    else:
        total_appe = (1.4 * academic_appe)

    user.total_appe = total_appe

    db.session.commit()
    return '', 200

@app.route('/delete_user/<int:user_id>', methods=['DELETE'])
def delete_user(user_id):
    user = User.query.get(user_id)
    if not user:
        return {"error": "User not found"}, 404
    # Delete all courses associated with the user
    Course.query.filter_by(user_id=user_id).delete()
    # Delete the user
    db.session.delete(user)
    db.session.commit()
    return {"message": "User and their courses deleted successfully"}, 200




@app.route('/generate_report', methods=['GET', 'POST'])
def generate_report():
    if request.method == 'POST':
        try:
            all_users = User.query.all()
            if not all_users:
                return jsonify({'error': 'No users found'}), 404

            # Prepare data for course sheet with user names
            course_data = []
            for u in all_users:
                name = u.first_name + " " + u.last_name
                user_courses = Course.query.filter_by(user_id=u.user_id).all()
                for course in user_courses:
                    course_data.append({
                        'Name': name,
                        'Course': course.course_name,
                        'Enroll': course.enroll,
                        'Didactic Credit': course.didactic_credit,
                        'Lab Credit': course.lab_credit,
                        'Coordinator': course.coordinator,
                        'Clinical Lead': course.clinical_lead,
                        'Lecture Total': course.lecture_total,
                        'Lab Total': course.lab_total,
                        'Lecture Faculty': course.lecture_faculty,
                        'Lab Design': course.lab_design,
                        'Lab Proctor': course.lab_proctor,
                        'Total': course.total
                    })
            df_courses = pd.DataFrame(course_data)

            # Prepare data for summary sheet
            summary_data = []
            teaching_data = []
            for u in all_users:
                name = u.first_name + " " + u.last_name
                user_courses = Course.query.filter_by(user_id=u.user_id).all()
                sum_of_courses = sum(course.total for course in user_courses)
                Total = sum_of_courses + (u.total_appe or 0) + (u.research_electives or 0) + (u.proctoring or 0) + (u.grading or 0)
                Total = round(Total, 2)
                Percentage = (Total * (0.7 / 21)) * 100
                Percentage = int(round(Percentage, 0))
                summary_data.append({
                    'Name': name,
                    'Sum of Courses': sum_of_courses,
                    'APPE (h)': u.total_appe,
                    'Research Electives': u.research_electives,
                    'Proctoring': u.proctoring,
                    'Grading': u.grading,
                    'Total sum of Courses': Total
                })
                teaching_data.append({
                    'Name': name,
                    'Teaching': Total,
                    '% Teaching': Percentage
                })
            df_summary1 = pd.DataFrame(summary_data)
            df_summary2 = pd.DataFrame(teaching_data)

            # Constant reference table data
            reference_data = {
                '%SH/Year': ['0 to 1', '2 to 4', '5 to 7', '8 to 10', '11 to 13', '14 to 16', '17 to 19', '20 to 22', '23 to 25', '26 to 28'],
                '% Teaching': [0, 10, 20, 30, 40, 50, 60, 70, 80, 90],
                'Current assumption' : ['2.75h per clinical APPE rotation', '1.4h per academic APPE rotation', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']
            }
            df_reference = pd.DataFrame(reference_data)

            # Create Excel file in memory with two sheets
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_courses.to_excel(writer, sheet_name='Courses', index=False)

                # Write first summary table
                df_summary1.to_excel(writer, sheet_name='Summary', startrow=0, index=False)

                # Write second summary table below the first with a gap of 2 rows
                startrow2 = len(df_summary1) + 3
                df_summary2.to_excel(writer, sheet_name='Summary', startrow=startrow2, index=False)

                # Write reference table below the second with a gap of 2 rows
                startrow3 = startrow2 + len(df_summary2) + 3
                df_reference.to_excel(writer, sheet_name='Summary', startrow=startrow3, index=False)

                ws_courses = writer.sheets['Courses']
                ws_summary = writer.sheets['Summary']

                def set_rounded_border(ws, min_row, max_row, min_col, max_col):
                    thin = Side(border_style="thin", color="000000")
                    thick = Side(border_style="thick", color="000000")

                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            cell = ws.cell(row=row, column=col)
                            border = Border()
                            # Top border
                            if row == min_row:
                                border = border + Border(top=thick)
                            else:
                                border = border + Border(top=thin)
                            # Bottom border
                            if row == max_row:
                                border = border + Border(bottom=thick)
                            else:
                                border = border + Border(bottom=thin)
                            # Left border
                            if col == min_col:
                                border = border + Border(left=thick)
                            else:
                                border = border + Border(left=thin)
                            # Right border
                            if col == max_col:
                                border = border + Border(right=thick)
                            else:
                                border = border + Border(right=thin)
                            cell.border = border

                def auto_adjust_column_width(ws):
                    for col in ws.columns:
                        max_length = 0
                        col_letter = get_column_letter(col[0].column)
                        for cell in col:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        adjusted_width = max_length + 2
                        ws.column_dimensions[col_letter].width = adjusted_width

                def merge_cells_in_column(ws, df, col_idx):
                    start_row = 2
                    last_value = None
                    merge_start = start_row
                    for row in range(start_row, len(df) + start_row):
                        cell_value = ws.cell(row=row, column=col_idx).value
                        if cell_value != last_value:
                            if row - 1 > merge_start:
                                ws.merge_cells(start_row=merge_start, start_column=col_idx, end_row=row - 1, end_column=col_idx)
                            merge_start = row
                            last_value = cell_value
                    if len(df) + start_row - 1 > merge_start:
                        ws.merge_cells(start_row=merge_start, start_column=col_idx, end_row=len(df) + start_row - 1, end_column=col_idx)

                merge_cells_in_column(ws_courses, df_courses, 1)
                merge_cells_in_column(ws_courses, df_courses, 2)

                set_rounded_border(ws_courses, 1, len(df_courses) + 1, 1, len(df_courses.columns))
                # Apply border separately for each summary table
                # First table
                set_rounded_border(ws_summary, 1, len(df_summary1) + 1, 1, len(df_summary1.columns))
                # Second table
                startrow2 = len(df_summary1) + 4
                set_rounded_border(ws_summary, startrow2, startrow2 + len(df_summary2), 1, len(df_summary2.columns))
                # Third table
                startrow3 = startrow2 + len(df_summary2) + 3
                set_rounded_border(ws_summary, startrow3, startrow3 + len(df_reference), 1, len(df_reference.columns))

                auto_adjust_column_width(ws_courses)
                auto_adjust_column_width(ws_summary)

            output.seek(0)
            return send_file(output, as_attachment=True, download_name='Teaching Percentage.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        except Exception as e:
            import traceback
            error_message = f"Error generating report: {str(e)}\n{traceback.format_exc()}"
            print(error_message)
            return jsonify({'error': error_message}), 500

    # else:  # GET method to download the file
    #     BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    #     reports_dir = os.path.join(BASE_DIR, '..', 'Frontend', 'Reports')
    #     os.makedirs(reports_dir, exist_ok=True)

    #     report_path = os.path.join(reports_dir, 'Teaching Percentage.xlsx')
    #     if not os.path.exists(report_path):
    #         return "Report not found", 404
    #     return send_file(report_path, as_attachment=True, download_name='Teaching Percentage.xlsx')


# Route to send survey email with link
@app.route('/send_survey_email/<int:user_id>', methods=['POST'])
def send_survey_email(user_id):
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'User not found'}), 404

    # Reset the survey access session key to allow resending survey
    session_key = f'survey_accessed_{user_id}'
    if session_key in session:
        session.pop(session_key)

    # Generate survey URL with ngrok public URL and https scheme
    survey_url = url_for('survey_form', user_id=user_id, _external=True, _scheme='https')

    msg = Message(subject="Please update your data",
                  sender=app.config['MAIL_USERNAME'],
                  recipients=[user.email])
    msg.body = f"Dear {user.first_name},\n\nPlease update your data by filling out the survey at the following link:\n{survey_url}\n\nThank you."
    try:
        mail.send(msg)
        return jsonify({'message': 'Survey email sent successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Route to render survey form with one-time password protection
@app.route('/survey/<int:user_id>', methods=['GET', 'POST'])
def survey_form(user_id):
    user = User.query.get(user_id)
    if not user:
        return "User not found", 404

    session_key = f'survey_accessed_{user_id}'

    if request.method == 'POST':
        if session.get(session_key):
            return "Password expired. You have already accessed the survey.", 403
        password = request.form.get('password')
        correct_password = os.environ.get("SURVEY_PASSWORD", "survey123")
        if password == correct_password:
            session[session_key] = True
            courses = Course.query.filter_by(user_id=user_id).all()
            return render_template('survey.html', user=user, courses=courses)
        else:
            error = "Invalid password. Please try again."
            return render_template('login.html', error=error, login_type='survey')
    else:
        if session.get(session_key):
            return "Password expired. You have already accessed the survey.", 403
        else:
            return render_template('login.html', login_type='survey')

# Route to receive survey submission
@app.route('/submit_survey/<int:user_id>', methods=['POST'])
def submit_survey(user_id):
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'User not found'}), 404

    data = request.form

    # Process course updates
    courses = Course.query.filter_by(user_id=user_id).all()
    for course in courses:
        enroll = data.get(f'enroll_{course.course_id}')
        coordinator = data.get(f'coordinator_{course.course_id}')
        clinical_lead = data.get(f'clinical_lead_{course.course_id}')
        lecture_faculty = data.get(f'lecture_faculty_{course.course_id}', type=int)
        lab_design = data.get(f'lab_design_{course.course_id}', type=float)
        lab_proctor = data.get(f'lab_proctor_{course.course_id}', type=float)
        if enroll or coordinator or clinical_lead or lecture_faculty or lab_design or lab_proctor:
            update_request = UpdateRequest(
                user_id=user_id,
                course_id=course.course_id,
                enroll=enroll,
                coordinator=coordinator,
                clinical_lead=clinical_lead,
                lecture_faculty=lecture_faculty,
                lab_design=lab_design,
                lab_proctor=lab_proctor,
                status='pending'
            )
            db.session.add(update_request)

    # Process APPE updates if any
    update_appe = data.get('update_appe')
    if update_appe == 'yes':
        clinical_appe = data.get('clinical_appe', type=int)
        academic_appe = data.get('academic_appe', type=int)
        update_request = UpdateRequest(
            user_id=user_id,
            clinical_appe=clinical_appe,
            academic_appe=academic_appe,
            status='pending'
        )
        db.session.add(update_request)

    db.session.commit()

    # Return a page that closes the window after submission
    return '''
    <html>
        <head>
            <script type="text/javascript">
                window.onload = function() {
                    window.close();
                };
            </script>
        </head>
        <body>
            <p>Thank you for submitting the survey. This window will close automatically.</p>
        </body>
    </html>
    '''

# Route to fetch pending updates for dashboard
@app.route('/pending_updates', methods=['GET'])
def pending_updates():
    updates = UpdateRequest.query.filter_by(status='pending').all()
    result = []
    for update in updates:
        user = User.query.get(update.user_id)
        # Fetch course info if enroll/coordinator/clinical_lead present
        course_info = None
        if update.enroll is not None or update.coordinator is not None or update.clinical_lead is not None:
            if update.course_id:
                course = Course.query.get(update.course_id)
                course_info = course.course_name if course else "Course update"
            else:
                course_info = "Course update"
        else:
            course_info = "APPE update"
        result.append({
            'id': update.id,
            'user_id': update.user_id,
            'user_name': f"{user.first_name} {user.last_name}",
            'course_info': course_info,
            'enroll': update.enroll,
            'coordinator': update.coordinator,
            'clinical_lead': update.clinical_lead,
            'lecture_faculty': update.lecture_faculty,
            'lab_design': update.lab_design,
            'lab_proctor': update.lab_proctor,
            'clinical_appe': update.clinical_appe,
            'academic_appe': update.academic_appe,
            'created_at': update.created_at.strftime('%Y-%m-%d %H:%M:%S')
        })
    return jsonify(result)

# Route to approve or reject update
@app.route('/update_request/<int:update_id>', methods=['POST'])
def update_request_action(update_id):
    action = request.json.get('action')
    update_request = UpdateRequest.query.get(update_id)
    if not update_request:
        return jsonify({'error': 'Update request not found'}), 404

    if action == 'approve':
        user = User.query.get(update_request.user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        # Update user fields based on update_request
        # Since enroll, coordinator, clinical_lead are course related, update only the specific course related to the update_request
        if update_request.enroll is not None or update_request.coordinator is not None or update_request.clinical_lead is not None or update_request.lecture_faculty is not None or update_request.lab_design is not None or update_request.lab_proctor is not None:
            if update_request.course_id:
                course = Course.query.get(update_request.course_id)
                if course:
                    if update_request.enroll is not None:
                        course.enroll = update_request.enroll
                    if update_request.coordinator is not None:
                        course.coordinator = update_request.coordinator
                    if update_request.clinical_lead is not None:
                        course.clinical_lead = update_request.clinical_lead
                    if update_request.lecture_faculty is not None:
                        course.lecture_faculty = update_request.lecture_faculty
                    if update_request.lab_design is not None:
                        course.lab_design = update_request.lab_design
                    if update_request.lab_proctor is not None:
                        course.lab_proctor = update_request.lab_proctor

                    # Recalculate total for the course using provided logic
                    enroll = course.enroll
                    didactic_credit = course.didactic_credit
                    lab_credit = course.lab_credit
                    coordinator = course.coordinator
                    clinical_lead = course.clinical_lead
                    lecture_total = course.lecture_total
                    lab_total = course.lab_total
                    lecture_faculty = course.lecture_faculty
                    lab_design = course.lab_design
                    lab_proctor = course.lab_proctor

                    if enroll and enroll.upper() == "Y":
                        part1 = didactic_credit if coordinator and coordinator.upper() == "Y" else 0
                    else:
                        part1 = 0

                    if coordinator and coordinator.upper() == "Y":
                        part2 = lab_credit if clinical_lead and clinical_lead.upper() == "NA" else 0
                    else:
                        part2 = 0

                    part3 = lab_credit if clinical_lead and clinical_lead.upper() == "Y" else 0
                    part4 = (lecture_faculty / lecture_total) * didactic_credit if lecture_total > 0 else 0
                    part5 = (((lab_design + lab_proctor) / lab_total) * lab_credit) if lab_total > 0 else 0

                    total = part1 + part2 + part3 + part4 + part5
                    total = round(total, 5)
                    course.total = total

                    db.session.commit()
            else:
                # fallback: update all courses of user
                courses = Course.query.filter_by(user_id=user.user_id).all()
                for course in courses:
                    if update_request.enroll is not None:
                        course.enroll = update_request.enroll
                    if update_request.coordinator is not None:
                        course.coordinator = update_request.coordinator
                    if update_request.clinical_lead is not None:
                        course.clinical_lead = update_request.clinical_lead
                    if update_request.lecture_faculty is not None:
                        course.lecture_faculty = update_request.lecture_faculty
                    if update_request.lab_design is not None:
                        course.lab_design = update_request.lab_design
                    if update_request.lab_proctor is not None:
                        course.lab_proctor = update_request.lab_proctor

                    # Recalculate total for each course using provided logic
                    enroll = course.enroll
                    didactic_credit = course.didactic_credit
                    lab_credit = course.lab_credit
                    coordinator = course.coordinator
                    clinical_lead = course.clinical_lead
                    lecture_total = course.lecture_total
                    lab_total = course.lab_total
                    lecture_faculty = course.lecture_faculty
                    lab_design = course.lab_design
                    lab_proctor = course.lab_proctor

                    if enroll and enroll.upper() == "Y":
                        part1 = didactic_credit if coordinator and coordinator.upper() == "Y" else 0
                    else:
                        part1 = 0

                    if coordinator and coordinator.upper() == "Y":
                        part2 = lab_credit if clinical_lead and clinical_lead.upper() == "NA" else 0
                    else:
                        part2 = 0

                    part3 = lab_credit if clinical_lead and clinical_lead.upper() == "Y" else 0
                    part4 = (lecture_faculty / lecture_total) * didactic_credit if lecture_total > 0 else 0
                    part5 = (((lab_design + lab_proctor) / lab_total) * lab_credit) if lab_total > 0 else 0

                    total = part1 + part2 + part3 + part4 + part5
                    total = round(total, 5)
                    course.total = total

                db.session.commit()

        if update_request.clinical_appe is not None and update_request.academic_appe is not None:
            user.clinical_appe = update_request.clinical_appe
            user.academic_appe = update_request.academic_appe

            if update_request.clinical_appe > 0 and update_request.academic_appe > 0:
                user.total_appe = (2.75 * update_request.clinical_appe) + (1.4 * update_request.academic_appe)
            elif update_request.clinical_appe > 0 and update_request.academic_appe == 0:
                user.total_appe = (2.75 * update_request.clinical_appe)
            else:
                user.total_appe = (1.4 * update_request.academic_appe)
            
        db.session.commit()
        update_request.status = 'approved'

        # Recalculate sum_of_courses, Total, and Percentage after approval
        user_courses = Course.query.filter_by(user_id=user.user_id).all()
        sum_of_courses = sum(c.total for c in user_courses)
        Total = sum_of_courses + (user.total_appe or 0) + (user.research_electives or 0) + (user.proctoring or 0) + (user.grading or 0)
        Total = round(Total, 2)
        Percentage = (Total * (0.7 / 21)) * 100
        Percentage = int(round(Percentage, 0))

        db.session.commit()
        return jsonify({
            'message': f'Update request {action}d successfully',
            'sum_of_courses': sum_of_courses,
            'Total': Total,
            'Percent': Percentage
        })
    elif action == 'reject':
        update_request.status = 'rejected'
    else:
        return jsonify({'error': 'Invalid action'}), 400

    db.session.commit()
    return jsonify({'message': f'Update request {action}d successfully'})

@app.route("/logout")
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('home'))

if __name__ == "__main__":
    with app.app_context():
        db_handler = DatabaseHandler(db_name=db_path)
        db_handler.setup_database()
    app.run()